"""
API endpoints для управления пользователями
"""
from fastapi import APIRouter, Depends, HTTPException, status, Query, Request, BackgroundTasks
from typing import Optional, List
from datetime import datetime, timedelta
import asyncio
import json

from ..auth.permissions import (
    RequireUserView, RequireUserEdit, get_current_active_user,
    log_admin_action
)
from ..models.user_models import (
    UserResponse, UsersListResponse, UserUpdateRequest,
    SubscriptionUpdateRequest, UserStatsResponse, BulkActionRequest,
    UserExportRequest, UserSearchRequest, UserFilterType,
    UserNotificationRequest, UserAnalyticsResponse
)
from database.universal_database import UniversalDatabase

router = APIRouter()


async def get_db(request: Request) -> UniversalDatabase:
    """Получить объект базы данных"""
    return request.state.db


@router.get("/", response_model=UsersListResponse)
@log_admin_action("view_users", "users")
async def get_users(
    page: int = Query(1, ge=1, description="Номер страницы"),
    per_page: int = Query(50, ge=1, le=200, description="Количество на странице"),
    search: Optional[str] = Query(None, max_length=100, description="Поиск по имени/username/ID"),
    filter_type: Optional[UserFilterType] = Query(None, description="Тип фильтра"),
    sort_by: str = Query("created_at", pattern="^(created_at|requests_used|last_request|username)$"),
    sort_order: str = Query("desc", pattern="^(asc|desc)$"),
    current_user = Depends(RequireUserView),
    db: UniversalDatabase = Depends(get_db)
):
    """Получить список пользователей с фильтрацией и пагинацией"""

    # Преобразуем filter_type в строку для совместимости с UniversalDatabase
    filter_str = None
    if filter_type:
        if filter_type == UserFilterType.subscribed:
            filter_str = "subscribed"
        elif filter_type == UserFilterType.unlimited:
            filter_str = "unlimited"
        elif filter_type == UserFilterType.blocked:
            filter_str = "blocked"
        elif filter_type == UserFilterType.bot_blocked:
            filter_str = "bot_blocked"
        elif filter_type == UserFilterType.active:
            filter_str = "active"

    # Получаем пользователей из базы данных
    result = await db.get_users_paginated(
        page=page,
        per_page=per_page,
        search=search,
        filter_type=filter_str
    )

    # Преобразуем в модели Pydantic
    users = []
    for user_data in result["users"]:
        user = UserResponse(**user_data)
        users.append(user)

    return UsersListResponse(
        users=users,
        total=result["total"],
        page=result["page"],
        per_page=result["per_page"],
        pages=result["pages"],
        filters_applied={
            "search": search,
            "filter_type": filter_type.value if filter_type else None,
            "sort_by": sort_by,
            "sort_order": sort_order
        }
    )


@router.get("/{user_id}", response_model=UserResponse)
@log_admin_action("view_user", "users")
async def get_user(
    user_id: int,
    current_user = Depends(RequireUserView),
    db: UniversalDatabase = Depends(get_db)
):
    """Получить информацию о конкретном пользователе"""
    user_data = await db.get_user(user_id)

    if not user_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Пользователь не найден"
        )

    return UserResponse(**user_data)


@router.delete("/{user_id}")
@log_admin_action("delete_user", "users")
async def delete_user(
    user_id: int,
    current_user = Depends(RequireUserEdit),
    db: UniversalDatabase = Depends(get_db)
):
    """Удалить пользователя"""
    # Проверяем существование пользователя
    user_data = await db.get_user(user_id)
    if not user_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Пользователь не найден"
        )

    # Удаляем пользователя
    success = await db.delete_user(user_id)

    if not success:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Ошибка при удалении пользователя"
        )

    return {"message": "Пользователь успешно удален", "user_id": user_id}


@router.put("/{user_id}", response_model=UserResponse)
@log_admin_action("update_user", "users")
async def update_user(
    user_id: int,
    update_data: UserUpdateRequest,
    current_user = Depends(RequireUserEdit),
    db: UniversalDatabase = Depends(get_db)
):
    """Обновить информацию о пользователе"""
    # Проверяем существование пользователя
    user_data = await db.get_user(user_id)
    if not user_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Пользователь не найден"
        )

    # Обновляем пользователя
    success = await db.update_user_permissions(
        user_id=user_id,
        unlimited_access=update_data.unlimited_access,
        blocked=update_data.blocked,
        notes=update_data.notes
    )

    if not success:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Не удалось обновить пользователя"
        )

    # Возвращаем обновленные данные
    updated_user = await db.get_user(user_id)
    return UserResponse(**updated_user)


# Удален дублирующийся endpoint - используется только manage_user_subscription ниже


@router.get("/{user_id}/stats", response_model=UserStatsResponse)
@log_admin_action("view_user_stats", "users")
async def get_user_stats(
    user_id: int,
    current_user = Depends(RequireUserView),
    db: UniversalDatabase = Depends(get_db)
):
    """Получить статистику пользователя"""
    # Проверяем существование пользователя
    user_data = await db.get_user(user_id)
    if not user_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Пользователь не найден"
        )

    # Получаем статистику запросов
    await db.adapter.connect()

    try:
        # Общее количество запросов
        if db.adapter.db_type == 'sqlite':
            query = "SELECT COUNT(*) as total FROM requests WHERE user_id = ?"
            params = (user_id,)
        else:  # PostgreSQL
            query = "SELECT COUNT(*) as total FROM requests WHERE user_id = $1"
            params = (user_id,)

        result = await db.adapter.fetch_one(query, params)
        total_requests = db._extract_count(result) if result else 0

        # Запросы за сегодня
        if db.adapter.db_type == 'sqlite':
            query = "SELECT COUNT(*) as today FROM requests WHERE user_id = ? AND DATE(created_at) = DATE('now')"
            params = (user_id,)
        else:  # PostgreSQL
            query = "SELECT COUNT(*) as today FROM requests WHERE user_id = $1 AND DATE(created_at) = CURRENT_DATE"
            params = (user_id,)

        result = await db.adapter.fetch_one(query, params)
        requests_today = db._extract_count(result) if result else 0

        # Запросы за неделю
        if db.adapter.db_type == 'sqlite':
            query = "SELECT COUNT(*) as week FROM requests WHERE user_id = ? AND created_at >= DATE('now', '-7 days')"
            params = (user_id,)
        else:  # PostgreSQL
            query = "SELECT COUNT(*) as week FROM requests WHERE user_id = $1 AND created_at >= CURRENT_DATE - INTERVAL '7 days'"
            params = (user_id,)

        result = await db.adapter.fetch_one(query, params)
        requests_week = db._extract_count(result) if result else 0

        # Запросы за месяц
        if db.adapter.db_type == 'sqlite':
            query = "SELECT COUNT(*) as month FROM requests WHERE user_id = ? AND created_at >= DATE('now', '-30 days')"
            params = (user_id,)
        else:  # PostgreSQL
            query = "SELECT COUNT(*) as month FROM requests WHERE user_id = $1 AND created_at >= CURRENT_DATE - INTERVAL '30 days'"
            params = (user_id,)

        result = await db.adapter.fetch_one(query, params)
        requests_month = db._extract_count(result) if result else 0

        # Первый и последний запрос
        if db.adapter.db_type == 'sqlite':
            query = "SELECT MIN(created_at) as first_request, MAX(created_at) as last_request FROM requests WHERE user_id = ?"
            params = (user_id,)
        else:  # PostgreSQL
            query = "SELECT MIN(created_at) as first_request, MAX(created_at) as last_request FROM requests WHERE user_id = $1"
            params = (user_id,)

        result = await db.adapter.fetch_one(query, params)
        request_dates = dict(result) if result else {"first_request": None, "last_request": None}

    finally:
        await db.adapter.disconnect()

    return UserStatsResponse(
        user_id=user_id,
        total_requests=total_requests,
        requests_today=requests_today,
        requests_week=requests_week,
        requests_month=requests_month,
        first_request=request_dates.get("first_request"),
        last_request=request_dates.get("last_request"),
        favorite_channels=[],  # TODO: Реализовать анализ популярных каналов
        subscription_history=[]  # TODO: Реализовать историю подписок
    )


@router.post("/{user_id}/subscription")
@log_admin_action("manage_subscription", "users")
async def manage_user_subscription(
    user_id: int,
    request: Request,
    current_user = Depends(RequireUserEdit),
    db: UniversalDatabase = Depends(get_db)
):
    """Управление подпиской пользователя"""
    # Проверяем существование пользователя
    user_data = await db.get_user(user_id)
    if not user_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Пользователь не найден"
        )

    try:
        body = await request.json()
        action = body.get('action')  # 'activate', 'extend' или 'cancel'
        duration_days = body.get('duration_days', 30)

        if action in ['activate', 'extend']:
            # Активируем или продлеваем подписку
            await db.adapter.connect()

            # Определяем новую дату окончания подписки
            if action == 'extend' and user_data.get('is_subscribed') and user_data.get('subscription_end'):
                # Продлеваем от текущей даты окончания
                current_end = user_data['subscription_end']
                if isinstance(current_end, str):
                    from datetime import datetime
                    current_end = datetime.fromisoformat(current_end.replace('Z', '+00:00'))

                # Если подписка еще активна, продлеваем от даты окончания
                if current_end > datetime.now():
                    new_end = current_end + timedelta(days=duration_days)
                else:
                    # Если подписка истекла, продлеваем от текущего момента
                    new_end = datetime.now() + timedelta(days=duration_days)
            else:
                # Активируем новую подписку от текущего момента
                new_end = datetime.now() + timedelta(days=duration_days)

            if db.adapter.db_type == 'sqlite':
                query = """
                    UPDATE users
                    SET is_subscribed = TRUE, subscription_end = ?
                    WHERE user_id = ?
                """
                params = (new_end, user_id)
            else:  # PostgreSQL
                query = """
                    UPDATE users
                    SET is_subscribed = TRUE, subscription_end = $1
                    WHERE user_id = $2
                """
                params = (new_end, user_id)

            await db.adapter.execute(query, params)
            await db.adapter.disconnect()

            action_text = "активирована" if action == 'activate' else "продлена"
            return {"success": True, "message": f"Подписка {action_text} на {duration_days} дней до {new_end.strftime('%d.%m.%Y')}"}

        elif action == 'cancel':
            # Отменяем подписку через прямой SQL запрос
            await db.adapter.connect()

            if db.adapter.db_type == 'sqlite':
                query = """
                    UPDATE users
                    SET is_subscribed = FALSE, subscription_end = NULL
                    WHERE user_id = ?
                """
                params = (user_id,)
            else:  # PostgreSQL
                query = """
                    UPDATE users
                    SET is_subscribed = FALSE, subscription_end = NULL
                    WHERE user_id = $1
                """
                params = (user_id,)

            await db.adapter.execute(query, params)
            await db.adapter.disconnect()

            return {"success": True, "message": "Подписка отменена"}
        else:
            return {"success": False, "message": "Неизвестное действие"}

    except Exception as e:
        logger.error(f"Ошибка управления подпиской пользователя {user_id}: {e}")
        return {"success": False, "message": "Внутренняя ошибка сервера"}


@router.get("/requests/{request_id}")
@log_admin_action("view_request_details", "requests")
async def get_request_details(
    request_id: int,
    current_user = Depends(RequireUserView),
    db: UniversalDatabase = Depends(get_db)
):
    """Получить детали конкретного запроса"""
    try:
        await db.adapter.connect()

        # Получаем детали запроса
        if db.adapter.db_type == 'sqlite':
            query = """
                SELECT r.*, u.username as user_username
                FROM requests r
                LEFT JOIN users u ON r.user_id = u.user_id
                WHERE r.id = ?
            """
            params = (request_id,)
        else:  # PostgreSQL
            query = """
                SELECT r.*, u.username as user_username
                FROM requests r
                LEFT JOIN users u ON r.user_id = u.user_id
                WHERE r.id = $1
            """
            params = (request_id,)

        result = await db.adapter.fetch_one(query, params)
        await db.adapter.disconnect()

        if not result:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Запрос не найден"
            )

        request_data = dict(result)

        # Парсим JSON поля если они есть
        if request_data.get('channels_input'):
            try:
                if isinstance(request_data['channels_input'], str):
                    request_data['channels_input'] = json.loads(request_data['channels_input'])
            except json.JSONDecodeError:
                request_data['channels_input'] = []

        if request_data.get('results'):
            try:
                if isinstance(request_data['results'], str):
                    request_data['results'] = json.loads(request_data['results'])
            except json.JSONDecodeError:
                request_data['results'] = []

        return request_data

    except Exception as e:
        logger.error(f"Ошибка получения деталей запроса {request_id}: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Ошибка при получении деталей запроса"
        )


@router.post("/bulk-action")
@log_admin_action("bulk_action", "users")
async def bulk_user_action(
    action_data: BulkActionRequest,
    current_user = Depends(RequireUserEdit),
    db: UniversalDatabase = Depends(get_db)
):
    """Массовые действия с пользователями"""
    results = {"success": [], "failed": []}

    for user_id in action_data.user_ids:
        try:
            if action_data.action == "block":
                success = await db.update_user_permissions(
                    user_id=user_id,
                    blocked=True,
                    notes=action_data.notes
                )
            elif action_data.action == "unblock":
                success = await db.update_user_permissions(
                    user_id=user_id,
                    blocked=False,
                    notes=action_data.notes
                )
            elif action_data.action == "grant_unlimited":
                success = await db.update_user_permissions(
                    user_id=user_id,
                    unlimited_access=True,
                    notes=action_data.notes
                )
            elif action_data.action == "revoke_unlimited":
                success = await db.update_user_permissions(
                    user_id=user_id,
                    unlimited_access=False,
                    notes=action_data.notes
                )
            elif action_data.action == "delete":
                success = await db.delete_user(user_id)
            else:
                success = False

            if success:
                results["success"].append(user_id)
            else:
                results["failed"].append(user_id)

        except Exception as e:
            results["failed"].append(user_id)

    return {
        "message": f"Обработано {len(action_data.user_ids)} пользователей",
        "results": results,
        "action": action_data.action
    }


@router.get("/search/", response_model=List[UserResponse])
@log_admin_action("search_users", "users")
async def search_users(
    query: str = Query(..., min_length=1, max_length=100),
    limit: int = Query(20, ge=1, le=100),
    current_user = Depends(RequireUserView),
    db: UniversalDatabase = Depends(get_db)
):
    """Поиск пользователей"""
    result = await db.get_users_paginated(
        page=1,
        per_page=limit,
        search=query
    )

    users = [UserResponse(**user_data) for user_data in result["users"]]
    return users


@router.get("/analytics/", response_model=UserAnalyticsResponse)
@log_admin_action("view_user_analytics", "users")
async def get_user_analytics(
    current_user = Depends(RequireUserView),
    db: UniversalDatabase = Depends(get_db)
):
    """Получить аналитику пользователей"""
    stats = await db.get_detailed_stats()
    user_stats = stats["users"]

    # Вычисляем сегменты
    total = user_stats.get("total", 0)
    segments = []

    if total > 0:
        # Используем корректные ключи из get_detailed_stats()
        active_subscribers = user_stats.get("subscribed", 0)
        unlimited_users = user_stats.get("unlimited", 0)  # Теперь есть в get_detailed_stats()
        blocked_users = user_stats.get("blocked", 0)
        new_week = user_stats.get("new_week", 0)  # TODO: добавить в get_detailed_stats()

        segments = [
            {
                "segment_name": "Активные подписчики",
                "count": active_subscribers,
                "percentage": round((active_subscribers / total) * 100, 2),
                "description": "Пользователи с активной подпиской"
            },
            {
                "segment_name": "Активные пользователи",
                "count": user_stats.get("active", 0),
                "percentage": round((user_stats.get("active", 0) / total) * 100, 2),
                "description": "Активные пользователи (не заблокированы)"
            },
            {
                "segment_name": "Заблокированные",
                "count": blocked_users,
                "percentage": round((blocked_users / total) * 100, 2),
                "description": "Заблокированные пользователи"
            }
        ]

        # Добавляем сегменты только если есть данные
        if unlimited_users > 0:
            segments.append({
                "segment_name": "Безлимитные пользователи",
                "count": unlimited_users,
                "percentage": round((unlimited_users / total) * 100, 2),
                "description": "Пользователи с безлимитным доступом"
            })

        if new_week > 0:
            segments.append({
                "segment_name": "Новые за неделю",
                "count": new_week,
                "percentage": round((new_week / total) * 100, 2),
                "description": "Зарегистрированы за последние 7 дней"
            })

    # Простой расчет роста (можно улучшить)
    growth_rate = 0.0
    new_month = user_stats.get("new_month", 0)
    if new_month > 0 and total > new_month:
        growth_rate = round((new_month / (total - new_month)) * 100, 2)

    return UserAnalyticsResponse(
        total_users=total,
        active_users=user_stats.get("active", 0),
        new_users_today=user_stats.get("new_today", 0),
        new_users_week=user_stats.get("new_week", 0),
        new_users_month=user_stats.get("new_month", 0),
        subscribers=user_stats.get("subscribed", 0),
        unlimited_users=user_stats.get("unlimited", 0),
        blocked_users=user_stats.get("blocked", 0),
        segments=segments,
        growth_rate=growth_rate,
        retention_rate=85.0  # TODO: Реализовать расчет retention rate
    )


@router.post("/export/")
@log_admin_action("export_users", "users")
async def export_users(
    export_data: UserExportRequest,
    background_tasks: BackgroundTasks,
    current_user = Depends(RequireUserView),
    db: UniversalDatabase = Depends(get_db)
):
    """Экспорт пользователей в CSV/Excel"""
    from fastapi.responses import StreamingResponse
    import io
    import csv

    # Получаем данные пользователей
    result = await db.get_users_paginated(
        page=1,
        per_page=10000,  # Большой лимит для экспорта
        search=export_data.search,
        filter_type=export_data.filter_type.value if export_data.filter_type else None
    )

    if export_data.format == "csv":
        # Создаем CSV
        output = io.StringIO()
        writer = csv.writer(output)

        # Заголовки
        headers = [
            "ID", "Username", "Имя", "Фамилия", "Дата регистрации",
            "Запросов использовано", "Подписка", "Дата окончания подписки",
            "Последний запрос", "Роль", "Безлимитный доступ", "Заблокирован", "Заметки"
        ]
        writer.writerow(headers)

        # Данные
        for user_data in result["users"]:
            user = UserResponse(**user_data)
            row = [
                user.user_id,
                user.username or "",
                user.first_name or "",
                user.last_name or "",
                user.created_at.strftime("%Y-%m-%d %H:%M:%S") if user.created_at else "",
                user.requests_used,
                "Да" if user.is_subscribed else "Нет",
                user.subscription_end.strftime("%Y-%m-%d %H:%M:%S") if user.subscription_end else "",
                user.last_request.strftime("%Y-%m-%d %H:%M:%S") if user.last_request else "",
                user.role,
                "Да" if user.unlimited_access else "Нет",
                "Да" if user.blocked else "Нет",
                user.notes or ""
            ]
            writer.writerow(row)

        output.seek(0)

        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
        )

    elif export_data.format == "xlsx":
        # Создаем Excel файл
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Пользователи"

            # Заголовки
            headers = [
                "ID", "Username", "Имя", "Фамилия", "Дата регистрации",
                "Запросов использовано", "Подписка", "Дата окончания подписки",
                "Последний запрос", "Роль", "Безлимитный доступ", "Заблокирован", "Заметки"
            ]

            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Данные
            for row_num, user_data in enumerate(result["users"], 2):
                user = UserResponse(**user_data)
                data = [
                    user.user_id,
                    user.username or "",
                    user.first_name or "",
                    user.last_name or "",
                    user.created_at.strftime("%Y-%m-%d %H:%M:%S") if user.created_at else "",
                    user.requests_used,
                    "Да" if user.is_subscribed else "Нет",
                    user.subscription_end.strftime("%Y-%m-%d %H:%M:%S") if user.subscription_end else "",
                    user.last_request.strftime("%Y-%m-%d %H:%M:%S") if user.last_request else "",
                    user.role,
                    "Да" if user.unlimited_access else "Нет",
                    "Да" if user.blocked else "Нет",
                    user.notes or ""
                ]

                for col, value in enumerate(data, 1):
                    ws.cell(row=row_num, column=col, value=value)

            # Автоширина колонок
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].auto_size = True

            # Сохраняем в память
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )

        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Библиотека openpyxl не установлена"
            )

    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Неподдерживаемый формат экспорта"
        )
